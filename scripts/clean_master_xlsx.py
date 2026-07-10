#!/usr/bin/env python3
"""
Rydder opp artist-linjer i en festival_data_master*.xlsx-fil FØR den importeres
med import_master_xlsx.py. Malen fylles med AI-generert research av varierende
kvalitet, og "Dag N"-cellene ender ofte opp med linjer som ikke er rene
artistnavn:

  1. Rene tall/årstall (f.eks. "1111") som havnet som egen linje.
  2. Lange semikolon-separerte oppføringer med klokkeslett/scene/dato bakt inn
     (f.eks. "18:15 Chabifönk Expérience; 20:00 Edensong; ...").
  3. Kommentar-/kilde-tekst ("Kilde: https://...") eller norske forklarende
     setninger som sneik seg inn som "artist".

Dette scriptet ENDRER IKKE originalfilen. Det skriver en ny kopi
("<navn>_renset.xlsx") der:
  - Rene tall og kommentar-/kilde-linjer er fjernet.
  - Semikolon-lister der klokkeslett/dato/scene kunne strippes trygt er
    splittet til én artist per linje.
  - Linjer scriptet er usikker på er latt urørt, men cellen er markert gul
    slik at du enkelt finner dem igjen i Excel og kan rydde manuelt.

Skriver også en tekstrapport (<navn>_rapport.txt) med alt som ble endret og
alt som ble flagget for manuell sjekk.

Run:  python3 scripts/clean_master_xlsx.py "/sti/til/festival_data_master....xlsx"
"""

import sys, re, os
import openpyxl
from openpyxl.styles import PatternFill

if len(sys.argv) < 2:
    print("Bruk: python3 scripts/clean_master_xlsx.py <sti-til-xlsx>", file=sys.stderr)
    sys.exit(1)

SRC = sys.argv[1]
BASE, EXT = os.path.splitext(SRC)
OUT_XLSX = f"{BASE}_renset{EXT}"
OUT_REPORT = f"{BASE}_rapport.txt"

NUMBER_RE = re.compile(r"^\s*\d{3,4}\s*$")
COMMENT_MARKERS = ("kilde:", "http://", "https://", "source:")

# Prosa-markører: dukker de opp i en linje uten noen ren artist, er linjen en
# forklarende kommentar ("ingen data funnet", "ikke publisert" osv.), ikke en
# artist. Da fjernes hele linjen (men logges alltid ordrett i rapporten).
PROSE_MARKERS = (
    "ikke publisert", "ikke funnet", "ikke verifisert", "ikke fylt",
    "ikke fordelt", "ikke tilgjengelig", "ikke utvidet", "ikke fullstendig",
    "ikke tydelig", "ingen 2026", "ingen trygg", "ingen glastonbury",
    "site åpner", "festivaldag", "pause officielle", "avsluttet etter",
    "programmet er lengre", "offisiell side", "offisiell timetable",
    "offisiell schedule", "offisiell programside", "dagsfordeling",
    "dagfordeling ikke", "arrangørens side", "neste festival",
    "prochain rendez-vous", "felt ikke", "programdetaljene", "lastes ikke",
    "for å unngå", "tilgjengelig tekstkilde", "tilgjengelig offisiell",
    "publisert i tilgjengelig", "ser ikke ut til", "artistfordeling per dag",
    "denne runden", "programme jour par jour",
)

TIME_TOK = r"\d{1,2}[:.]\d{2}"
DATE_PREFIX_RE = re.compile(r"^\d{4}-\d{2}-\d{2}(?:[–—-]\d{1,2})?\s*(?:" + TIME_TOK + r"[^:]*)?:\s*")
ALPHA_PREFIX_RE = re.compile(r"^([^;:0-9\n]{2,40}):\s+(.+)$")
LEADING_TIME_RE = re.compile(r"^(?:" + TIME_TOK + r"(?:\s*[&–—-]\s*" + TIME_TOK + r")*)\s+")
TRAILING_PAREN_TIME_RE = re.compile(r"\s*\([^)]*" + TIME_TOK + r"[^)]*\)\s*$")
TIME_ANYWHERE_RE = re.compile(TIME_TOK)
HAS_LETTER_RE = re.compile(r"[^\W\d_]", re.UNICODE)
# Tre eller flere komma-separerte navn med stor forbokstav = en artistliste
# gjemt i prosa ("... inkluderer Iron Maiden, Bring Me The Horizon, Limp Bizkit").
# Slike linjer skal IKKE fjernes selv om de har prosa-markører – de flagges.
_CAP = r"[A-ZÀ-ÖØ-Þ][\w'’&.\-]*(?:\s+[A-ZÀ-ÖØ-Þ0-9][\w'’&.\-]*){0,4}"
COMMA_NAME_LIST_RE = re.compile(_CAP + r"\s*,\s*" + _CAP + r"\s*,\s*" + _CAP)
YELLOW = PatternFill("solid", fgColor="FFF2CC")


def is_pure_number(line):
    return bool(NUMBER_RE.match(line))


def is_comment(line):
    low = line.lower()
    return any(m in low for m in COMMENT_MARKERS)


def has_prose_marker(s):
    low = s.lower()
    return any(m in low for m in PROSE_MARKERS)


def looks_like_artist(seg):
    """En streng som trygt kan stå som artistnavn: ingen klokkeslett, spillested-
    tankestrek, kolon (scene/verk-prefiks) eller forklarende prosa."""
    seg = seg.strip()
    if not (2 <= len(seg) <= 55):
        return False
    if not HAS_LETTER_RE.search(seg):
        return False
    if TIME_ANYWHERE_RE.search(seg) or ":" in seg:
        return False
    if " – " in seg or " — " in seg:   # spillested-separator (en/em-dash)
        return False
    if "http" in seg.lower() or has_prose_marker(seg):
        return False
    return True


def strip_segment(seg):
    """Fjerner ledende klokkeslett ("20:00 X") og etterstilt tids-parentes
    ("X (20:50 & 22:45)"). Returnerer (renset, hadde_ledende_tid)."""
    seg = seg.strip()
    seg = TRAILING_PAREN_TIME_RE.sub("", seg)
    had_leading_time = bool(LEADING_TIME_RE.match(seg))
    seg = LEADING_TIME_RE.sub("", seg)
    return seg.strip(" -–—:•\t."), had_leading_time


def try_split(line):
    """Splitter en semikolon-liste til rene artistnavn hvis ALLE segmenter blir
    trygge artistnavn. Ellers None (linjen flagges for manuell gjennomgang).

    Dato-timeplaner ("2026-08-20: 17:00 X; 21:00 Y") med ledende klokkeslett og
    uten scene-prefiks er klassisk/teater-program der segmentene ofte er verk
    eller spillesteder, ikke artister – de flagges bevisst i stedet for å
    splittes."""
    body = line.strip()
    had_date_prefix = bool(DATE_PREFIX_RE.match(body))
    body = DATE_PREFIX_RE.sub("", body)

    alpha = ALPHA_PREFIX_RE.match(body)
    has_alpha_prefix = bool(alpha)
    if has_alpha_prefix:
        body = alpha.group(2)

    if ";" not in body:
        return None
    raw = [p.strip() for p in body.split(";") if p.strip()]
    if len(raw) < 2:
        return None

    cleaned = []
    any_leading_time = False
    for p in raw:
        c, had_time = strip_segment(p)
        any_leading_time = any_leading_time or had_time
        cleaned.append(c)

    # Dato-timeplan med klokkeslett per punkt og ingen scene-prefiks: for risikabelt.
    if any_leading_time and had_date_prefix and not has_alpha_prefix:
        return None
    if not all(looks_like_artist(c) for c in cleaned):
        return None
    return cleaned


def artist_segments(line):
    """Antall semikolon-segmenter i linjen som ser ut som ekte artistnavn.
    Brukes til å avgjøre om en prosa-linje har noe artistinnhold verdt å beholde."""
    body = DATE_PREFIX_RE.sub("", line.strip())
    alpha = ALPHA_PREFIX_RE.match(body)
    if alpha:
        body = alpha.group(2)
    segs = [strip_segment(p)[0] for p in body.split(";")]
    return [s for s in segs if looks_like_artist(s)]


def main():
    print(f"Leser {SRC} ...")
    wb = openpyxl.load_workbook(SRC)
    ws = wb["Festivaler"]

    header = [c.value for c in ws[1]]
    day_col_idxs = [i + 1 for i, name in enumerate(header) if name and str(name).startswith("Dag")]
    festival_col = header.index("Festival") + 1

    # Nullstill ALLE fargemarkeringer i arket først, slik at de eneste gule
    # cellene i resultatfila er de dette scriptet flagger for gjennomgang.
    no_fill = PatternFill(fill_type=None)
    cleared = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.fill_type is not None:
                cell.fill = no_fill
                cleared += 1

    report = []
    removed_total = 0
    split_total = 0
    flagged_total = 0
    festivals_touched = set()

    for row in ws.iter_rows(min_row=2):
        fest_name = row[festival_col - 1].value
        if not fest_name:
            continue
        for col_idx in day_col_idxs:
            cell = row[col_idx - 1]
            raw = cell.value
            if raw is None:
                continue
            lines = [ln.strip() for ln in str(raw).splitlines()]
            lines = [ln for ln in lines if ln]
            if not lines:
                continue

            new_lines = []
            cell_changed = False
            cell_flagged = False
            local_report = []

            for ln in lines:
                if is_pure_number(ln):
                    local_report.append(f"    FJERNET (rent tall): {ln!r}")
                    removed_total += 1
                    cell_changed = True
                    continue
                if is_comment(ln):
                    local_report.append(f"    FJERNET (kilde/kommentar): {ln!r}")
                    removed_total += 1
                    cell_changed = True
                    continue
                # Ren prosa/«ingen data»-kommentar uten noe artistinnhold å redde.
                # (Har linjen en enumerert artistliste, semikolon- eller komma-
                # separert, beholdes den og flagges i stedet for å fjernes.)
                if (has_prose_marker(ln) and not artist_segments(ln)
                        and not COMMA_NAME_LIST_RE.search(ln)):
                    local_report.append(f"    FJERNET (ingen artist-data): {ln!r}")
                    removed_total += 1
                    cell_changed = True
                    continue
                split_result = try_split(ln) if ";" in ln else None
                if split_result:
                    local_report.append(f"    SPLITTET: {ln!r}")
                    for piece in split_result:
                        local_report.append(f"        -> {piece!r}")
                    new_lines.extend(split_result)
                    split_total += 1
                    cell_changed = True
                    continue
                # Alt annet som ser rotete ut (lang liste, prosa med navn gjemt
                # inni, klokkeslett) beholdes urørt, men flagges for manuell sjekk.
                suspicious = (
                    (len(ln) > 55 and (";" in ln or TIME_ANYWHERE_RE.search(ln)))
                    or has_prose_marker(ln)
                    or (len(ln) > 55 and COMMA_NAME_LIST_RE.search(ln))
                )
                if suspicious:
                    local_report.append(f"    MANUELL SJEKK (usikker splitting): {ln!r}")
                    flagged_total += 1
                    cell_flagged = True
                    new_lines.append(ln)
                    continue
                new_lines.append(ln)

            if cell_changed or cell_flagged:
                festivals_touched.add(fest_name)
                report.append(f"\n## {fest_name} — {header[col_idx - 1]} (rad {cell.row})")
                report.extend(local_report)

            if cell_changed:
                cell.value = "\n".join(new_lines) if new_lines else None
            if cell_flagged:
                cell.fill = YELLOW

    wb.save(OUT_XLSX)
    with open(OUT_REPORT, "w") as f:
        f.write(f"Ryddet {SRC}\n")
        f.write(f"{len(festivals_touched)} festivaler berørt | "
                f"{removed_total} linjer fjernet | {split_total} linjer splittet | "
                f"{flagged_total} linjer flagget for manuell sjekk\n")
        f.write("\n".join(report))
        f.write("\n")

    print(f"\nFjernet {cleared} gamle fargemarkeringer (nullstilt til ingen farge)")
    print(f"{len(festivals_touched)} festivaler berørt")
    print(f"{removed_total} linjer fjernet (rene tall / kilde-tekst)")
    print(f"{split_total} linjer splittet til flere artister")
    print(f"{flagged_total} linjer flagget for manuell sjekk (cellen er gulmarkert i Excel)")
    print(f"\nSkrev {OUT_XLSX}")
    print(f"Skrev {OUT_REPORT}")


if __name__ == "__main__":
    main()
