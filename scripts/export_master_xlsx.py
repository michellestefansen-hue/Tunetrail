#!/usr/bin/env python3
"""
Exports everything in the database to a friendly, editable Excel master file.

One sheet, "Festivaler": one row per festival edition. The first column is the
festival's slug — the key the importer matches on. It is greyed out because
changing it points the row at a different festival (or at none). Every other
column, the name included, is safe to edit: correcting a name here renames the
festival rather than detaching the row from it.

Then one column per day ("Dag 1", "Dag 2", ...). Each day cell starts with the
date on its own line, followed by that day's artists, one per line:

    2026-08-06
    Saxon
    Judas Priest

"Dag 1" is the festival's first programme day, "Dag 2" the next, and so on (they
follow the real dates, so festivals with gaps stay compact). The "Program" column
shows at a glance who is missing a line-up (red "MANGLER").

Edit this file by hand, then run scripts/import_master_xlsx.py <file> to turn it
back into SQL you can paste into the Supabase SQL editor.

Run:  python3 scripts/export_master_xlsx.py
Writes: festival_data_master.xlsx (in the current directory)
"""

import json, urllib.request, urllib.parse
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SB_URL = "https://espxguvjupinrxobyabz.supabase.co"
SB_KEY = "sb_publishable_2RC59nLrzNoKDVvgu_oo2g_Ut7c6V-Z"
SELECT = ("id,name,slug,website_url,image_url,city,region,venue_name,country,latitude,longitude,"
          "category,festival_editions(year,date_from,date_to,ticket_url,program)")

print("Fetching from database...")
req = urllib.request.Request(
    f"{SB_URL}/rest/v1/festivals?select={urllib.parse.quote(SELECT, safe=',()')}&limit=5000",
    headers={"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"},
)
festivals = json.load(urllib.request.urlopen(req))
festivals.sort(key=lambda f: (f.get("country") or "", f.get("name") or ""))
print(f"  {len(festivals)} festivals")


def to_date(s):
    try:
        return date.fromisoformat(str(s)[:10])
    except (ValueError, TypeError):
        return None


def edition_rows():
    for f in festivals:
        for ed in (f.get("festival_editions") or [None]):
            yield f, ed


def program_days(ed):
    """Days that actually have artists, sorted by date."""
    days = [d for d in (ed.get("program") or []) if (d.get("artists") or [])]
    return sorted(days, key=lambda d: str(d.get("date") or ""))


# --- Pass 1: how many day columns we need (max days with a line-up) --------
n_day_cols = max((len(program_days(ed)) for f, ed in edition_rows() if ed), default=0)
n_day_cols = max(n_day_cols, 1)

# --- Build the sheet -------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Festivaler"

HEAD_FILL = PatternFill("solid", fgColor="FF4E50")
HEAD_FONT = Font(bold=True, color="FFFFFF")
MISSING_FILL = PatternFill("solid", fgColor="FFD6D6")   # soft red
MISSING_FONT = Font(bold=True, color="C0392B")
KEY_FILL = PatternFill("solid", fgColor="F1EFE8")       # muted: "leave this alone"
KEY_FONT = Font(color="8A8178")
DAY_ALIGN = Alignment(vertical="top", wrap_text=True)

base_cols = ["Slug", "Festival", "Land", "By", "Region", "Sted/venue", "Sjanger", "Nettside",
             "Bilde-URL", "Breddegrad", "Lengdegrad",
             "Billett-URL", "Årstall", "Fra dato", "Til dato", "Program"]
day_cols = [f"Dag {i + 1}" for i in range(n_day_cols)]
cols = base_cols + day_cols
ws.append(cols)
PROG_COL = base_cols.index("Program") + 1
FIRST_DAY_COL = len(base_cols) + 1

n_missing = 0
for f, ed in edition_rows():
    ed = ed or {}
    days = program_days(ed)
    n_art = sum(len(d.get("artists") or []) for d in days)
    missing = n_art == 0
    status = "MANGLER" if missing else f"{len(days)} dager · {n_art} artister"
    if missing:
        n_missing += 1

    # Each Dag column: date on the first line, then the artists.
    day_cells = [None] * n_day_cols
    for i, day in enumerate(days[:n_day_cols]):
        artists = [a.get("name") for a in (day.get("artists") or []) if a.get("name")]
        lines = [str(day.get("date") or "")] + artists
        day_cells[i] = "\n".join(lines)

    ws.append([
        f.get("slug"), f.get("name"), f.get("country"), f.get("city"), f.get("region"),
        f.get("venue_name"), f.get("category"), f.get("website_url"), f.get("image_url"),
        f.get("latitude"), f.get("longitude"),
        ed.get("ticket_url"), ed.get("year"), ed.get("date_from"), ed.get("date_to"),
        status, *day_cells,
    ])
    row = ws.max_row
    # The slug is the key the importer matches on — greyed out to read as
    # "don't touch". Everything else, including the name, is safe to edit.
    key = ws.cell(row=row, column=1)
    key.fill, key.font = KEY_FILL, KEY_FONT
    if missing:
        c = ws.cell(row=row, column=PROG_COL)
        c.fill, c.font = MISSING_FILL, MISSING_FONT
    for j in range(n_day_cols):
        ws.cell(row=row, column=FIRST_DAY_COL + j).alignment = DAY_ALIGN

# Header styling + freeze + filter.
for c in range(1, len(cols) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill, cell.font = HEAD_FILL, HEAD_FONT
    cell.alignment = Alignment(vertical="center")
ws.freeze_panes = "C2"  # keep Slug + Festival visible while scrolling
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

widths = [26, 30, 14, 16, 14, 24, 20, 30, 34, 12, 12, 30, 8, 12, 12, 20] + [26] * n_day_cols
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

out = "festival_data_master.xlsx"
wb.save(out)
print(f"  {n_day_cols} day columns · {n_missing} festivals missing a programme")
print(f"Wrote {out}")
