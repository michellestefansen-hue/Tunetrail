#!/usr/bin/env python3
"""
Imports the Excel day-program workbook into the new festival_editions structure.

Reads:
  - "Dag-for-dag 2026"  -> day structure (date, weekday) per festival
  - "Artister per dag"  -> artists with stage + time per day
  - "Festivalprogrammer"-> festival-level info (country, venue, website, ticket, genre)

Matches each festival to an existing DB festival by name; unmatched festivals
with a real program are created as new rows (geocoded via Nominatim).

Writes supabase/seed_excel_editions.sql (festivals inserts + editions upserts).
Run:  python3 scripts/import_excel_editions.py <path-to-xlsx>
"""

import sys, re, json, time, unicodedata, urllib.request, urllib.parse
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else \
    "/Users/michellestefansen/Downloads/festivalprogrammer_europa_2026_dagprogram_utfylt_del12.xlsx"
SB_URL = "https://espxguvjupinrxobyabz.supabase.co"
SB_KEY = "sb_publishable_2RC59nLrzNoKDVvgu_oo2g_Ut7c6V-Z"
UA = "Tunetrail/1.0 (festival map; github.com/michellestefansen-hue/Tunetrail)"

COUNTRY_EN = {
    "Norge": "Norway", "Sverige": "Sweden", "Danmark": "Denmark", "Finland": "Finland",
    "Island": "Iceland", "Storbritannia": "United Kingdom", "Irland": "Ireland",
    "Nederland": "Netherlands", "Belgia": "Belgium", "Tyskland": "Germany",
    "Frankrike": "France", "Spania": "Spain", "Portugal": "Portugal", "Italia": "Italy",
    "Sveits": "Switzerland", "Østerrike": "Austria", "Polen": "Poland", "Tsjekkia": "Czechia",
    "Ungarn": "Hungary", "Kroatia": "Croatia", "Romania": "Romania", "Hellas": "Greece",
    "Slovenia": "Slovenia", "Slovakia": "Slovakia", "Estland": "Estonia", "Latvia": "Latvia",
    "Litauen": "Lithuania", "Serbia": "Serbia", "Bulgaria": "Bulgaria",
}

def norm(s):
    s = (s or "").lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"\b(festival|festivalen|fest|open air|openair)\b", " ", s)
    return re.sub(r"[^a-z0-9]+", " ", s).strip()

def slugify(s):
    s = s.lower().replace("&", " og ")
    for a, b in [("ø","o"),("æ","ae"),("å","a"),("ö","o"),("ä","a"),("ü","u"),("ß","ss")]:
        s = s.replace(a, b)
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"(^-+|-+$)", "", re.sub(r"[^a-z0-9]+", "-", s))[:60]

def category(genre):
    g = (genre or "").lower()
    if re.search(r"techno|house|trance", g): return "Techno & House"
    if "metal" in g: return "Metal"
    if re.search(r"punk|hardcore", g): return "Punk & Hardcore"
    if "indie" in g: return "Indie"
    if re.search(r"elektron|edm|electronic|dance", g): return "Elektronisk & Dans"
    if re.search(r"hip-?hop|rap|r&b", g): return "Hip-Hop & R&B"
    if "rock" in g: return "Rock & Alternativ"
    if "pop" in g: return "Pop & Mainstream"
    if re.search(r"jazz|blues|soul|funk", g): return "Jazz & Soul"
    if re.search(r"klassisk|classical|opera", g): return "Klassisk"
    if re.search(r"folk|country|americana|vise", g): return "Folk & Americana"
    if re.search(r"reggae|world|latin|ska", g): return "Reggae & World"
    return "Blandet/Flersjanger"

def d(v):
    return v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else None

print("Reading workbook...", file=sys.stderr)
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

# Festival-level info keyed by festival name
prog = {}
ws = wb["Festivalprogrammer"]
rows = list(ws.iter_rows(values_only=True))
h = {name: i for i, name in enumerate(rows[0])}
for r in rows[1:]:
    name = r[h["Festival"]]
    if not name:
        continue
    prog[name] = {
        "country": r[h["Land"]],
        "venue": r[h["Sted/venue"]],
        "website": r[h["Offisiell nettside"]],
        "ticket": r[h["Billettkjøp URL"]],
        "genre": r[h["Hovedsjanger"]],
    }

# Day structure keyed by festival id
ws = wb["Dag-for-dag 2026"]
rows = list(ws.iter_rows(values_only=True))
h = {name: i for i, name in enumerate(rows[0])}
festivals = {}  # fid -> {name, days: {date: {label, artists: []}}}
for r in rows[1:]:
    fid = r[h["Festival ID"]]
    date = d(r[h["Dato"]])
    if not fid or not date:
        continue
    f = festivals.setdefault(fid, {"name": r[h["Festival"]], "days": {}})
    f["days"].setdefault(date, {"label": r[h["Ukedag"]], "artists": []})

# Detailed artists keyed by (festival id, date)
ws = wb["Artister per dag"]
rows = list(ws.iter_rows(values_only=True))
h = {name: i for i, name in enumerate(rows[0])}
for r in rows[1:]:
    fid = r[h["Festival ID"]]
    date = d(r[h["Dato"]])
    artist = r[h["Artist"]]
    if not fid or not date or not artist:
        continue
    f = festivals.get(fid)
    if not f:
        continue
    day = f["days"].setdefault(date, {"label": r[h["Ukedag"]], "artists": []})
    day["artists"].append({"name": str(artist).strip(),
                            "stage": (r[h["Stage"]] or None),
                            "time": (str(r[h["Tid"]]).strip() if r[h["Tid"]] else None)})

# DB festivals for matching
print("Fetching DB festivals...", file=sys.stderr)
req = urllib.request.Request(f"{SB_URL}/rest/v1/festivals?select=name,slug&limit=5000",
                             headers={"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"})
db = json.load(urllib.request.urlopen(req))
db_by_norm = {}
for row in db:
    db_by_norm.setdefault(norm(row["name"]), row["slug"])

def match_slug(name):
    n = norm(name)
    if not n:
        return None
    if n in db_by_norm:
        return db_by_norm[n]
    if len(n) >= 5:
        for dn, slug in db_by_norm.items():
            if len(dn) >= 5 and (n in dn or dn in n):
                return slug
    return None

def geocode(query):
    url = f"https://nominatim.openstreetmap.org/search?format=json&limit=1&q={urllib.parse.quote(query)}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": UA})
        res = json.load(urllib.request.urlopen(req, timeout=15))
        if res:
            return float(res[0]["lon"]), float(res[0]["lat"])
    except Exception as e:
        print(f"  geocode failed for {query}: {e}", file=sys.stderr)
    return None

# Build editions + collect new festivals to create
new_festivals = []  # dicts
editions = []       # dicts
used_new_slugs = set()

for fid, f in festivals.items():
    name = f["name"]
    info = prog.get(name, {})
    days = sorted(f["days"].keys())
    if not days:
        continue
    program = []
    for date in days:
        day = f["days"][date]
        # dedupe artists by name+stage
        seen = set()
        artists = []
        for a in day["artists"]:
            key = (a["name"].lower(), (a["stage"] or ""))
            if key not in seen:
                seen.add(key)
                artists.append(a)
        program.append({"date": date, "day_label": day["label"], "artists": artists})

    slug = match_slug(name)
    if not slug:
        slug = slugify(name)
        base = slug
        n = 2
        while slug in used_new_slugs:
            slug = f"{base}-{n}"; n += 1
        used_new_slugs.add(slug)
        # geocode
        venue = (info.get("venue") or "").split("/")[0].split(",")[0].strip()
        country = info.get("country") or ""
        q = ", ".join(x for x in [venue, COUNTRY_EN.get(country, country)] if x)
        coord = geocode(q) if q else None
        time.sleep(1.1)  # Nominatim politeness
        if not coord:
            print(f"  SKIP new (no coords): {name}", file=sys.stderr)
            continue
        new_festivals.append({
            "slug": slug, "name": name, "country": country,
            "city": venue or None, "venue": info.get("venue"),
            "website": info.get("website"), "category": category(info.get("genre")),
            "lng": coord[0], "lat": coord[1],
        })

    editions.append({
        "slug": slug, "year": int(days[0][:4]),
        "date_from": days[0], "date_to": days[-1],
        "ticket_url": info.get("ticket"), "website": info.get("website"),
        "program": program,
    })

print(f"\nEditions: {len(editions)} | new festivals: {len(new_festivals)}", file=sys.stderr)

def sql_str(v):
    return "null" if v is None or v == "" else "'" + str(v).replace("'", "''") + "'"

out = ["-- Auto-generated by scripts/import_excel_editions.py",
       f"-- {len(editions)} editions, {len(new_festivals)} new festivals\n"]

if new_festivals:
    out.append("insert into festivals (slug, name, city, country, venue_name, website_url, category, latitude, longitude, source, external_id) values")
    out.append(",\n".join(
        f"  ({sql_str(f['slug'])}, {sql_str(f['name'])}, {sql_str(f['city'])}, {sql_str(f['country'])}, "
        f"{sql_str(f['venue'])}, {sql_str(f['website'])}, {sql_str(f['category'])}, {f['lat']}, {f['lng']}, "
        f"'excel', {sql_str(f['slug'])})"
        for f in new_festivals))
    out.append("on conflict (slug) do nothing;\n")

# Backfill website on matched festivals where missing
web_rows = [e for e in editions if e["website"]]
if web_rows:
    out.append("update festivals fest set website_url = coalesce(fest.website_url, v.website)")
    out.append("from (values")
    out.append(",\n".join(f"  ({sql_str(e['slug'])}, {sql_str(e['website'])})" for e in web_rows))
    out.append(") as v(slug, website) where fest.slug = v.slug;\n")

out.append("insert into festival_editions (festival_id, year, date_from, date_to, ticket_url, program, source)")
out.append("select fest.id, v.year, v.date_from::date, v.date_to::date, v.ticket_url, v.program::jsonb, 'excel'")
out.append("from (values")
vals = []
for e in editions:
    program_json = json.dumps(e["program"], ensure_ascii=False)
    vals.append(f"  ({sql_str(e['slug'])}, {e['year']}, {sql_str(e['date_from'])}, {sql_str(e['date_to'])}, "
                f"{sql_str(e['ticket_url'])}, {sql_str(program_json)})")
out.append(",\n".join(vals))
out.append(") as v(slug, year, date_from, date_to, ticket_url, program)")
out.append("join festivals fest on fest.slug = v.slug")
out.append("""on conflict (festival_id, year) do update set
  date_from = excluded.date_from,
  date_to = excluded.date_to,
  ticket_url = coalesce(excluded.ticket_url, festival_editions.ticket_url),
  program = excluded.program,
  source = 'excel',
  updated_at = now();""")

open("supabase/seed_excel_editions.sql", "w").write("\n".join(out))
print("Wrote supabase/seed_excel_editions.sql", file=sys.stderr)
